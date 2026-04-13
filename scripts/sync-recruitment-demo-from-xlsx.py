#!/usr/bin/env python3
"""
Regenerate src/data/recruitment-demo-data.ts from recruitment_financial_model.xlsx.

Usage:
  python3 scripts/sync-recruitment-demo-from-xlsx.py /path/to/recruitment_financial_model.xlsx

Month labels in the workbook (e.g. Jan-23 … Dec-24) are shifted +1 year for display (Jan 24 … Dec 25).

Requires: pip install openpyxl
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "data" / "recruitment-demo-data.ts"


def shift_month_label(lbl: str) -> str:
    s = str(lbl).strip().replace("-", " ")
    m = re.match(r"^([A-Za-z]{3})\s+(\d{2})$", s)
    if not m:
        return s
    mon, yy = m.group(1).title()[:3], int(m.group(2))
    return f"{mon} {yy + 1:02d}"


def num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return 0.0


def find_row(ws, label: str) -> int | None:
    target = label.strip().lower()
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if c1 is None:
            continue
        if str(c1).strip().lower() == target:
            return r
    return None


def month_slice(ws, row: int) -> list[float]:
    return [num(ws.cell(row, c).value) for c in range(3, 27)]


def emit_arr(name: str, values: list, int_vals: bool = False) -> str:
    if int_vals:
        body = ",\n  ".join(str(int(round(x))) for x in values)
    else:
        body = ",\n  ".join(str(int(x)) if x == int(x) else str(round(x, 4)) for x in values)
    return f"export const {name} = [\n  {body}\n];\n\n"


def main() -> None:
    xlsx = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else Path.home() / "Downloads" / "recruitment_financial_model.xlsx"
    if not xlsx.exists():
        raise SystemExit(f"File not found: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    kpi = wb["KPI Dashboard"]
    pl = wb["P&L"]
    cf = wb["Cash Flow"]

    hdr = [kpi.cell(3, c).value for c in range(3, 27)]
    months = [shift_month_label(h) for h in hdr]

    def need(ws, label: str) -> list[float]:
        r = find_row(ws, label)
        if not r:
            raise SystemExit(f"Missing row {label!r} in {ws.title}")
        return month_slice(ws, r)

    nfi = need(kpi, "Total Net Fee Income (NFI)")
    perm_fee = need(kpi, "Perm Fee Income")
    contract_margin = need(kpi, "Contract Margin Income")
    total_placements = need(kpi, "Total Placements (Perm + New Cont.)")
    perm_placements = need(kpi, "Perm Placements Made")
    interviews = need(kpi, "Interviews Conducted")
    conversion = need(kpi, "Placement Conversion Rate")
    avg_perm_fee = need(kpi, "Avg Perm Fee per Placement")
    rev_per_consultant = need(kpi, "Revenue per Consultant")
    staff_pct_nfi = need(kpi, "Staff Costs as % of NFI")
    ebitda_margin_kpi = need(kpi, "EBITDA Margin %")
    commission_paid = need(kpi, "Total Commission Paid")
    new_contract_starts = need(kpi, "New Contract Starts")
    time_to_fill = need(kpi, "Avg Time to Fill - Perm (days)")

    staff_costs = need(pl, "Total Staff Costs")
    total_opex = need(pl, "Total Operating Expenses")
    ebitda = need(pl, "EBITDA")
    operating_profit = need(pl, "Operating Profit (EBIT)")
    corp_tax = need(pl, "Corporation Tax (25%)")
    pat = need(pl, "Profit After Tax (PAT)")

    rent = need(pl, "Office Rent (Serviced)")
    job_boards = need(pl, "Job Boards & Advertising")
    crm = need(pl, "CRM & Tech Stack")
    bd = need(pl, "BD & Marketing")
    acct = need(pl, "Accountancy & Compliance")
    ins = need(pl, "Insurance (PI & Employer)")
    phone = need(pl, "Phone & Communications")
    bank = need(pl, "Bank Charges & Misc")

    cash_in = need(cf, "Total Cash Inflows")
    cash_out = need(cf, "Total Cash Outflows")
    net_cash = need(cf, "Net Cash Movement")
    closing = need(cf, "Closing Cash Balance")

    n = len(months)
    assert all(len(x) == n for x in [nfi, staff_costs, total_opex, ebitda])

    gross_profit = [nfi[i] - staff_costs[i] for i in range(n)]
    gross_margin_pct = [100.0 * gross_profit[i] / nfi[i] if nfi[i] else 0.0 for i in range(n)]

    mom_growth = [0.0]
    for i in range(1, n):
        p, c = nfi[i - 1], nfi[i]
        mom_growth.append(round(((c - p) / p) * 100, 4) if p else 0.0)

    churned = [0] * n
    monthly_churn_pct = [0.0] * n
    nrr = [100.0 + (i % 5) * 0.12 for i in range(n)]

    cac_list: list[float] = []
    ltv_list: list[float] = []
    for i in range(n):
        acq_spend = job_boards[i] + bd[i]
        pp = max(int(perm_placements[i]), 1)
        cac_list.append(round(acq_spend / pp, 2))
        ltv_list.append(round(avg_perm_fee[i] * 2.2, 2))

    ltv_cac = [round(ltv_list[i] / cac_list[i], 1) if cac_list[i] else 0.0 for i in range(n)]
    cac_payback = [
        round(cac_list[i] / max(nfi[i] / max(total_placements[i], 1), 0.01), 2) for i in range(n)
    ]

    blocks: list[str] = []
    blocks.append(
        "/**\n * Recruitment agency demo — generated from recruitment_financial_model.xlsx\n"
        f" * Source: {xlsx.name}\n"
        " * Re-run: python3 scripts/sync-recruitment-demo-from-xlsx.py \"<path-to-xlsx>\"\n"
        " */\n\n"
    )
    blocks.append("export const months = [\n  " + ",\n  ".join(f'"{m}"' for m in months) + "\n];\n\n")
    blocks.append(emit_arr("revenue", nfi))
    blocks.append(emit_arr("totalCogs", staff_costs))
    blocks.append(emit_arr("grossProfit", gross_profit))
    blocks.append(emit_arr("grossMarginPct", gross_margin_pct))
    blocks.append(emit_arr("totalOpex", total_opex))
    blocks.append(emit_arr("ebitda", ebitda))
    blocks.append(emit_arr("operatingProfit", operating_profit))
    blocks.append(emit_arr("corpTax", corp_tax))
    blocks.append(emit_arr("cashInflows", cash_in))
    blocks.append(emit_arr("cashOutflows", cash_out))
    blocks.append(emit_arr("netCashMovement", net_cash))
    blocks.append(emit_arr("closingCash", closing))
    blocks.append(emit_arr("mrr", nfi))
    blocks.append("export const arr = mrr.map((m) => m * 12);\n\n")
    blocks.append(emit_arr("momGrowthPct", mom_growth))
    blocks.append(emit_arr("customers", total_placements, int_vals=True))
    blocks.append(emit_arr("newCustomers", new_contract_starts, int_vals=True))
    blocks.append(emit_arr("churnedCustomers", churned, int_vals=True))
    blocks.append(emit_arr("monthlyChurnPct", monthly_churn_pct))
    blocks.append(emit_arr("nrr", nrr))
    blocks.append(emit_arr("cac", cac_list))
    blocks.append(emit_arr("ltv", ltv_list))
    blocks.append(emit_arr("ltvCacRatio", ltv_cac))
    blocks.append(emit_arr("cacPayback", cac_payback))

    def emit_line(name: str, arr: list[float]) -> str:
        body = ",\n    ".join(str(int(x)) if x == int(x) else str(round(x, 4)) for x in arr)
        return f"    {name}: [\n      {body}\n    ],"

    ra_lines = [
        "  permFeeIncome: permFeeIncomeRecruitment,",
        "  contractMarginIncome: contractMarginIncomeRecruitment,",
        "  totalPlacements: totalPlacementsRecruitment,",
        "  permPlacements: permPlacementsRecruitment,",
        "  interviewsConducted: interviewsConductedRecruitment,",
        "  placementConversionRate: placementConversionRateRecruitment,",
        "  avgPermFeePerPlacement: avgPermFeePerPlacementRecruitment,",
        "  revenuePerConsultant: revenuePerConsultantRecruitment,",
        "  staffCostPctNfi: staffCostPctNfiRecruitment,",
        "  ebitdaMarginPct: ebitdaMarginPctRecruitment,",
        "  totalCommissionPaid: totalCommissionPaidRecruitment,",
        "  avgTimeToFillDays: avgTimeToFillDaysRecruitment,",
        "  recruitmentOpex: {",
        f"    {emit_line('officeRent', rent)}",
        f"    {emit_line('jobBoardsAdvertising', job_boards)}",
        f"    {emit_line('crmTech', crm)}",
        f"    {emit_line('bdMarketing', bd)}",
        f"    {emit_line('accountancyCompliance', acct)}",
        f"    {emit_line('insurance', ins)}",
        f"    {emit_line('phoneComms', phone)}",
        f"    {emit_line('bankMisc', bank)}",
        "  },",
    ]

    blocks.append(emit_arr("permFeeIncomeRecruitment", perm_fee))
    blocks.append(emit_arr("contractMarginIncomeRecruitment", contract_margin))
    blocks.append(emit_arr("totalPlacementsRecruitment", total_placements, int_vals=True))
    blocks.append(emit_arr("permPlacementsRecruitment", perm_placements, int_vals=True))
    blocks.append(emit_arr("interviewsConductedRecruitment", interviews, int_vals=True))
    blocks.append(emit_arr("placementConversionRateRecruitment", conversion))
    blocks.append(emit_arr("avgPermFeePerPlacementRecruitment", avg_perm_fee))
    blocks.append(emit_arr("revenuePerConsultantRecruitment", rev_per_consultant))
    blocks.append(emit_arr("staffCostPctNfiRecruitment", staff_pct_nfi))
    blocks.append(emit_arr("ebitdaMarginPctRecruitment", ebitda_margin_kpi))
    blocks.append(emit_arr("totalCommissionPaidRecruitment", commission_paid))
    blocks.append(emit_arr("avgTimeToFillDaysRecruitment", time_to_fill, int_vals=True))

    blocks.append("export const recruitmentAgency = {\n" + "\n".join(ra_lines) + "\n};\n\n")
    blocks.append('export const financialYearEnd = new Date("2026-06-30");\n')
    blocks.append('export const nextVatReturnDueDate = new Date("2026-11-07");\n')
    blocks.append("export const corpTaxRate = 0.25;\n")

    OUT.write_text("".join(blocks), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
