#!/usr/bin/env python3
"""
Regenerate src/data/barbershop-demo-data.ts from barbershop_financial_model.xlsx.

Usage:
  python3 scripts/sync-barbershop-demo-from-xlsx.py /path/to/barbershop_financial_model.xlsx

Requires: pip install openpyxl
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "data" / "barbershop-demo-data.ts"


def header_row(ws) -> tuple:
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row and row[0] == "Line Item" and row[1] == "Notes":
            return row
    return None


def month_labels_from_header(ws) -> list[str]:
    row = header_row(ws)
    if not row:
        raise SystemExit("Could not find header row with 'Line Item', 'Notes'")
    labels: list[str] = []
    for c in range(2, 26):
        v = row[c] if c < len(row) else None
        if v is None:
            break
        s = str(v).strip().replace("-", " ")
        labels.append(s)
    if len(labels) != 24:
        raise SystemExit(f"Expected 24 month columns, got {len(labels)}: {labels}")
    return labels


def row_by_label(ws, label: str) -> tuple | None:
    for row in ws.iter_rows(min_row=1, max_row=500, values_only=True):
        if row and row[0] == label:
            return row
    return None


def need(ws, label: str) -> tuple:
    r = row_by_label(ws, label)
    if not r:
        raise SystemExit(f"Missing row: {label!r}")
    return r


def month_slice(row: tuple) -> list:
    return [row[i] if i < len(row) else None for i in range(2, 26)]


def nums(vals: list) -> list[float]:
    out = []
    for v in vals:
        if v is None:
            out.append(0.0)
        elif isinstance(v, (int, float)):
            out.append(float(v))
        else:
            out.append(0.0)
    return out


def fmt_arr_int(name: str, values: list[int], indent=0) -> str:
    pad = " " * indent
    body = ",\n  ".join(str(x) for x in values)
    return f"{pad}export const {name} = [\n  {body}\n{pad}];\n"


def emit_float_arr(name: str, values: list[float]) -> str:
    lines = ",\n  ".join(str(int(x)) if x == int(x) else str(round(x, 4)) for x in values)
    return f"export const {name} = [\n  {lines}\n];\n\n"


def emit_shop_operating_costs(costs: dict[str, list[float]]) -> str:
    keys = [
        "rentRates",
        "utilities",
        "consumables",
        "insurance",
        "cleaning",
        "marketing",
        "professionalAndBank",
    ]
    parts = ["export const shopOperatingCosts = {"]
    for k in keys:
        arr = costs[k]
        body = ",\n    ".join(str(int(round(x))) for x in arr)
        parts.append(f"  {k}: [\n    {body}\n  ],")
    parts.append("};\n\n")
    return "\n".join(parts)


def derive_new_and_churn(customers: list[int]) -> tuple[list[int], list[int]]:
    new_c = [0] * len(customers)
    churned = [0] * len(customers)
    for i in range(1, len(customers)):
        d = customers[i] - customers[i - 1]
        if d >= 0:
            new_c[i] = int(d)
        else:
            churned[i] = int(-d)
    return new_c, churned


def main() -> None:
    xlsx = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else Path.home() / "Downloads" / "barbershop_financial_model.xlsx"
    if not xlsx.exists():
        print(f"File not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    pl = wb["P&L"]
    cf = wb["Cash Flow"]
    kpi = wb["KPI Dashboard"]

    months = month_labels_from_header(pl)

    revenue = nums(month_slice(need(pl, "Total Revenue")))
    owner_service = nums(month_slice(need(pl, "Owner Service Revenue (ex-VAT)")))
    chair_rent = nums(month_slice(need(pl, "Chair Rent Income")))
    consumables = nums(month_slice(need(pl, "Consumables (Blades, Products)")))
    total_op_cost = nums(month_slice(need(pl, "Total Operating Costs")))
    total_cogs = consumables
    total_opex = [max(0.0, total_op_cost[i] - consumables[i]) for i in range(24)]

    rent_rates = nums(month_slice(need(pl, "Shop Rent & Business Rates")))
    shop_utilities = nums(month_slice(need(pl, "Utilities (Gas, Electric, Water)")))
    shop_insurance = nums(month_slice(need(pl, "Public Liability Insurance")))
    shop_cleaning = nums(month_slice(need(pl, "Cleaning & Sundries")))
    shop_marketing = nums(month_slice(need(pl, "Marketing & Advertising")))
    shop_acct = nums(month_slice(need(pl, "Accountancy & Bookkeeping")))
    shop_bank = nums(month_slice(need(pl, "Bank Charges")))
    prof_bank = [shop_acct[i] + shop_bank[i] for i in range(24)]

    gross_profit = [revenue[i] - total_cogs[i] for i in range(24)]
    gross_margin_pct = [round((gross_profit[i] / revenue[i]) * 100, 2) if revenue[i] else 0.0 for i in range(24)]

    ebitda = nums(month_slice(need(pl, "EBITDA")))
    operating_profit = nums(month_slice(need(pl, "Operating Profit (EBIT)")))

    # Sole trader demo — no corporation tax in model
    corp_tax = [0] * 24

    cash_in = nums(month_slice(need(cf, "Total Cash Inflows")))
    cash_out = nums(month_slice(need(cf, "Total Cash Outflows")))
    net_mov = nums(month_slice(need(cf, "Net Cash Movement")))
    closing = nums(month_slice(need(cf, "Closing Cash Balance")))

    mrr = nums(month_slice(need(kpi, "Total Revenue (ex-VAT)")))
    revenue_per_chair_day = nums(month_slice(need(kpi, "Revenue per Chair per Day (£)")))
    mom_growth = [0.0]
    for i in range(1, 24):
        p, c = mrr[i - 1], mrr[i]
        mom_growth.append(round(((c - p) / p) * 100, 4) if p else 0.0)

    customers = [int(round(x)) for x in nums(month_slice(need(kpi, "Total Est. Customers (Monthly)")))]
    new_customers, churned_customers = derive_new_and_churn(customers)
    monthly_churn = [0.0]
    for i in range(1, 24):
        prev = customers[i - 1]
        monthly_churn.append(round((churned_customers[i] / prev) * 100, 4) if prev else 0.0)

    nrr = [round(100.0 + (i % 5) * 0.35, 2) for i in range(24)]
    cac = [int(42 + (i % 7) * 4) for i in range(24)]
    arpu = [mrr[i] / max(customers[i], 1) for i in range(24)]
    ltv = [int(round(arpu[i] * 18)) for i in range(24)]
    ltv_cac = [round(ltv[i] / max(cac[i], 1), 1) for i in range(24)]
    cac_payback = [round(cac[i] / max(arpu[i], 0.01), 2) for i in range(24)]

    wb.close()

    blocks: list[str] = []
    blocks.append(
        "/**\n * Barbershop demo — generated from barbershop_financial_model.xlsx\n"
        f" * Source: {xlsx.name}\n"
        " * Re-run: python3 scripts/sync-barbershop-demo-from-xlsx.py \"<path-to-xlsx>\"\n"
        " */\n\n"
    )
    blocks.append("export const months = [\n  " + ",\n  ".join(f'"{m}"' for m in months) + "\n];\n\n")
    blocks.append(emit_float_arr("revenue", revenue))
    blocks.append(emit_float_arr("ownerServiceRevenue", owner_service))
    blocks.append(emit_float_arr("chairRentRevenue", chair_rent))
    blocks.append(emit_float_arr("totalCogs", total_cogs))
    blocks.append(emit_float_arr("grossProfit", gross_profit))
    blocks.append(emit_float_arr("grossMarginPct", gross_margin_pct))
    blocks.append(emit_float_arr("totalOpex", total_opex))
    blocks.append(emit_float_arr("ebitda", ebitda))
    blocks.append(emit_float_arr("operatingProfit", operating_profit))
    blocks.append(fmt_arr_int("corpTax", corp_tax))
    blocks.append("\n")
    blocks.append("export const cashInflows = [\n  " + ",\n  ".join(str(int(x)) for x in cash_in) + "\n];\n\n")
    blocks.append("export const cashOutflows = [\n  " + ",\n  ".join(str(int(x)) for x in cash_out) + "\n];\n\n")
    blocks.append("export const netCashMovement = [\n  " + ",\n  ".join(str(int(x)) for x in net_mov) + "\n];\n\n")
    blocks.append("export const closingCash = [\n  " + ",\n  ".join(str(int(x)) for x in closing) + "\n];\n\n")
    blocks.append(emit_float_arr("mrr", mrr))
    blocks.append("export const arr = mrr.map((m) => m * 12);\n\n")
    blocks.append(emit_float_arr("momGrowthPct", mom_growth))
    blocks.append(fmt_arr_int("customers", customers))
    blocks.append(fmt_arr_int("newCustomers", new_customers))
    blocks.append(fmt_arr_int("churnedCustomers", churned_customers))
    blocks.append(emit_float_arr("monthlyChurnPct", monthly_churn))
    blocks.append(emit_float_arr("nrr", nrr))
    blocks.append(fmt_arr_int("cac", cac))
    blocks.append(fmt_arr_int("ltv", ltv))
    blocks.append(emit_float_arr("ltvCacRatio", ltv_cac))
    blocks.append(emit_float_arr("cacPayback", cac_payback))
    blocks.append(emit_float_arr("revenuePerChairPerDay", revenue_per_chair_day))
    blocks.append(
        emit_shop_operating_costs(
            {
                "rentRates": rent_rates,
                "utilities": shop_utilities,
                "consumables": consumables,
                "insurance": shop_insurance,
                "cleaning": shop_cleaning,
                "marketing": shop_marketing,
                "professionalAndBank": prof_bank,
            }
        )
    )
    blocks.append("export const seatCount = 6;\n\n")
    blocks.append('export const financialYearEnd = new Date("2025-12-31");\n')
    blocks.append('export const nextVatReturnDueDate = new Date("2026-08-07");\n')
    blocks.append("export const corpTaxRate = 0;\n")

    OUT.write_text("".join(blocks), encoding="utf-8")
    print(f"Wrote {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
